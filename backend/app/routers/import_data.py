from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Any
from datetime import datetime
import csv
import io
from urllib.parse import quote

from openpyxl import Workbook, load_workbook

from app.database import get_db
from app.models import Customer, Lead
from app.routers.utils import require_user

router = APIRouter()

CUSTOMER_FIELDS = [
    {"key": "name", "label": "客户名称", "required": True, "aliases": ["客户名称", "客户名", "name"]},
    {"key": "industry", "label": "行业", "required": False, "aliases": ["行业", "industry"]},
    {"key": "address", "label": "地址", "required": False, "aliases": ["地址", "address"]},
    {"key": "website", "label": "网址", "required": False, "aliases": ["网址", "网站", "website"]},
    {"key": "level", "label": "客户等级", "required": False, "aliases": ["客户等级", "等级", "level"]},
    {"key": "description", "label": "描述", "required": False, "aliases": ["描述", "备注", "description"]},
]

LEAD_FIELDS = [
    {"key": "name", "label": "线索名称", "required": True, "aliases": ["线索名称", "线索名", "name"]},
    {"key": "company", "label": "公司名称", "required": False, "aliases": ["公司名称", "公司", "company"]},
    {"key": "contact_name", "label": "联系人", "required": False, "aliases": ["联系人", "contact_name"]},
    {"key": "contact_phone", "label": "联系电话", "required": False, "aliases": ["联系电话", "电话", "手机", "contact_phone"]},
    {"key": "source", "label": "来源", "required": False, "aliases": ["来源", "source"]},
    {"key": "quality", "label": "质量", "required": False, "aliases": ["质量", "quality"]},
    {"key": "industry", "label": "行业", "required": False, "aliases": ["行业", "industry"]},
    {"key": "notes", "label": "备注", "required": False, "aliases": ["备注", "说明", "notes"]},
]

TEMPLATES = {
    "customers": {
        "sheet": "客户导入",
        "fields": CUSTOMER_FIELDS,
        "sample": ["示例客户有限公司", "金融", "北京市海淀区", "https://example.com", "A", "重点客户"],
    },
    "leads": {
        "sheet": "线索导入",
        "fields": LEAD_FIELDS,
        "sample": ["等保建设项目", "示例集团", "张三", "13800000000", "官网", "热", "政数", "预算已确认"],
    },
}

LEAD_SOURCE_MAP = {
    "官网": "website",
    "网站": "website",
    "website": "website",
    "展会": "exhibition",
    "exhibition": "exhibition",
    "伙伴引荐": "partner",
    "伙伴": "partner",
    "partner": "partner",
    "电话": "phone",
    "phone": "phone",
    "招标网站": "bidding",
    "招标": "bidding",
    "bidding": "bidding",
    "其他": "other",
    "other": "other",
}

LEAD_QUALITY_MAP = {
    "冷": "cold",
    "cold": "cold",
    "温": "warm",
    "warm": "warm",
    "热": "hot",
    "hot": "hot",
}


class ImportConfirmReq(BaseModel):
    rows: list[dict[str, Any]]
    duplicate_strategy: str = "skip"


def _template_config(type_: str) -> dict[str, Any]:
    cfg = TEMPLATES.get(type_)
    if not cfg:
        raise HTTPException(400, "不支持的导入类型")
    return cfg


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _alias_map(fields: list[dict[str, Any]]) -> dict[str, str]:
    mapping = {}
    for field in fields:
        mapping[field["key"].lower()] = field["key"]
        for alias in field["aliases"]:
            mapping[alias.strip().lower()] = field["key"]
    return mapping


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(400, "文件没有表头")
    return [{k: _clean(v) for k, v in row.items()} for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件没有可导入的数据")
    headers = [_clean(v) for v in rows[0]]
    if not any(headers):
        raise HTTPException(400, "文件没有表头")
    parsed = []
    for values in rows[1:]:
        if not any(_clean(v) for v in values):
            continue
        parsed.append({headers[i]: _clean(values[i]) if i < len(values) else "" for i in range(len(headers)) if headers[i]})
    return parsed


def _parse_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    if lower.endswith(".csv"):
        return _parse_csv(content)
    raise HTTPException(400, "仅支持 .xlsx 或 .csv 文件")


def _normalize_rows(type_: str, raw_rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    cfg = _template_config(type_)
    aliases = _alias_map(cfg["fields"])
    normalized = []
    for raw in raw_rows:
        item = {}
        for source_key, value in raw.items():
            if source_key is None:
                continue
            key = aliases.get(str(source_key).strip().lower())
            if key:
                item[key] = _clean(value)
        for field in cfg["fields"]:
            item.setdefault(field["key"], "")
        if type_ == "leads":
            item["source"] = LEAD_SOURCE_MAP.get(item.get("source", "").lower(), LEAD_SOURCE_MAP.get(item.get("source", ""), item.get("source", "")))
            item["quality"] = LEAD_QUALITY_MAP.get(item.get("quality", "").lower(), LEAD_QUALITY_MAP.get(item.get("quality", ""), item.get("quality", "")))
        normalized.append(item)
    return normalized


def _unique_key(type_: str, row: dict[str, str]) -> str:
    if type_ == "customers":
        return row.get("name", "").strip().lower()
    return "|".join([
        row.get("name", "").strip().lower(),
        row.get("company", "").strip().lower(),
        row.get("contact_phone", "").strip(),
    ])


def _exists_in_db(type_: str, row: dict[str, str], db: Session) -> bool:
    if type_ == "customers":
        return bool(db.query(Customer).filter(Customer.name == row.get("name", "")).first())
    q = db.query(Lead).filter(Lead.name == row.get("name", ""))
    if row.get("company"):
        q = q.filter(Lead.company == row["company"])
    if row.get("contact_phone"):
        q = q.filter(Lead.contact_phone == row["contact_phone"])
    return bool(q.first())


def _preview_rows(type_: str, rows: list[dict[str, str]], db: Session) -> dict[str, Any]:
    cfg = _template_config(type_)
    seen: set[str] = set()
    output = []
    errors = []
    for idx, row in enumerate(rows, start=2):
        item = dict(row)
        item["_row"] = idx
        row_errors = []
        for field in cfg["fields"]:
            if field["required"] and not item.get(field["key"]):
                row_errors.append(f"{field['label']}不能为空")
        unique = _unique_key(type_, item)
        if unique and unique in seen:
            row_errors.append("文件内重复数据")
        elif unique:
            seen.add(unique)
        if unique and _exists_in_db(type_, item, db):
            item["_duplicate"] = True
            item["_warnings"] = "数据库中已存在，确认导入时将跳过"
        item["_valid"] = len(row_errors) == 0
        item["_errors"] = "；".join(row_errors)
        if row_errors:
            errors.append(f"第 {idx} 行：{item['_errors']}")
        output.append(item)
    valid_count = sum(1 for row in output if row["_valid"])
    error_count = len(output) - valid_count
    duplicate_count = sum(1 for row in output if row.get("_duplicate"))
    return {
        "headers": [field["key"] for field in cfg["fields"]],
        "rows": output,
        "total": len(output),
        "valid_count": valid_count,
        "error_count": error_count,
        "duplicate_count": duplicate_count,
        "errors": errors,
    }


def _xlsx_response(type_: str) -> StreamingResponse:
    cfg = _template_config(type_)
    wb = Workbook()
    ws = wb.active
    ws.title = cfg["sheet"]
    ws.append([field["label"] for field in cfg["fields"]])
    ws.append(cfg["sample"])
    ws.freeze_panes = "A2"
    for idx, field in enumerate(cfg["fields"], start=1):
        ws.column_dimensions[chr(64 + idx)].width = max(14, len(field["label"]) + 8)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "客户导入模板.xlsx" if type_ == "customers" else "线索导入模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.get("/template")
def get_template(type: str = Query("customers")):
    return _xlsx_response(type)


@router.get("/template/{type}")
def get_template_legacy(type: str):
    return _xlsx_response(type)


@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    type: str = "customers",
    db: Session = Depends(get_db),
    user=Depends(require_user),
):
    content = await file.read()
    raw_rows = _parse_upload(file.filename or "", content)
    rows = _normalize_rows(type, raw_rows)
    return _preview_rows(type, rows, db)


@router.post("/confirm")
async def confirm_import(
    request: Request,
    type: str = "customers",
    db: Session = Depends(get_db),
    user=Depends(require_user),
):
    rows: list[dict[str, str]]
    duplicate_strategy = "skip"
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        payload = ImportConfirmReq.model_validate(await request.json())
        rows = _normalize_rows(type, payload.rows)
        duplicate_strategy = payload.duplicate_strategy
    elif "multipart/form-data" in content_type:
        form = await request.form()
        file = form.get("file")
        if file is None or not hasattr(file, "read"):
            raise HTTPException(400, "缺少导入文件")
        content = await file.read()
        rows = _normalize_rows(type, _parse_upload(file.filename or "", content))
    else:
        raise HTTPException(400, "缺少导入数据")

    preview = _preview_rows(type, rows, db)
    valid_rows = [row for row in preview["rows"] if row["_valid"]]
    imported = 0
    skipped = 0
    for row in valid_rows:
        if _exists_in_db(type, row, db):
            skipped += 1
            continue
        if duplicate_strategy != "skip":
            raise HTTPException(400, "当前仅支持 skip 重复策略")
        if type == "customers":
            db.add(Customer(
                name=row.get("name", ""),
                industry=row.get("industry") or None,
                address=row.get("address") or None,
                website=row.get("website") or None,
                level=row.get("level") or "C",
                description=row.get("description") or None,
                owner_id=user.id,
            ))
        elif type == "leads":
            db.add(Lead(
                name=row.get("name", ""),
                company=row.get("company") or None,
                contact_name=row.get("contact_name") or None,
                contact_phone=row.get("contact_phone") or None,
                source=row.get("source") or "other",
                quality=row.get("quality") or "cold",
                industry=row.get("industry") or None,
                notes=row.get("notes") or None,
                assigned_to=user.id,
            ))
        else:
            raise HTTPException(400, "不支持的导入类型")
        imported += 1
    db.commit()
    return {
        "message": f"成功导入 {imported} 条，跳过 {skipped} 条",
        "imported": imported,
        "skipped": skipped,
        "errors": preview["errors"],
        "error_count": preview["error_count"],
    }
